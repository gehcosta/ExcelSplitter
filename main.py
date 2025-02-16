import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

def dividir_contatos(arquivo_entrada, pasta_saida, nome_base, tamanho_lote, progress_var):
    # Criar pasta de saída com o mesmo nome do arquivo base
    pasta_saida = os.path.join(pasta_saida, nome_base)
    
    if os.path.exists(pasta_saida):
        messagebox.showerror("Erro", "A pasta com o nome especificado já existe. Escolha outro nome.")
        return
    
    os.makedirs(pasta_saida, exist_ok=True)
    
    # Ler o arquivo Excel
    df = pd.read_excel(arquivo_entrada, usecols=["ID", "Contato", "Telefone"], engine="openpyxl")
    
    # Definir o tamanho do lote
    total_contatos = len(df)
    num_arquivos = (total_contatos // tamanho_lote) + (1 if total_contatos % tamanho_lote else 0)
    
    # Inverter a ordem dos dados para que os primeiros IDs fiquem no primeiro arquivo
    df = df.iloc[::-1].reset_index(drop=True)
    
    for i in range(num_arquivos):
        inicio = i * tamanho_lote
        fim = inicio + tamanho_lote
        df_lote = df.iloc[inicio:fim]
        
        nome_arquivo = os.path.join(pasta_saida, f"{nome_base} - {i + 1}.xlsx")
        df_lote.to_excel(nome_arquivo, index=False, engine="openpyxl")
        
        # Remover formatação em negrito e bordas do cabeçalho
        wb = load_workbook(nome_arquivo)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=False)
            cell.border = Border(left=Side(border_style=None),
                                 right=Side(border_style=None),
                                 top=Side(border_style=None),
                                 bottom=Side(border_style=None))
        wb.save(nome_arquivo)
        
        progress_var.set(int(((i + 1) / num_arquivos) * 100))
        root.update_idletasks()
    
    messagebox.showinfo("Concluído", f"Processo concluído! {num_arquivos} arquivos foram gerados.")

def selecionar_arquivo():
    caminho = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Excel files", "*.xlsx")])
    entrada_var.set(caminho)

def selecionar_pasta():
    caminho = filedialog.askdirectory(title="Selecione a pasta onde deseja salvar os arquivos")
    saida_var.set(caminho)

def iniciar_divisao(event=None):
    arquivo_entrada = entrada_var.get()
    pasta_saida = saida_var.get()
    nome_base = nome_var.get().strip()
    tamanho_lote = tamanho_lote_var.get().strip()
    
    if not arquivo_entrada or not pasta_saida or not nome_base or not tamanho_lote.isdigit():
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos corretamente!")
        return
    
    dividir_contatos(arquivo_entrada, pasta_saida, nome_base, int(tamanho_lote), progress_var)

# Criar interface gráfica
root = tk.Tk()
root.title("Divisor de Contatos Excel")
root.geometry("500x350")
root.configure(bg="#f0f0f0")

frame = tk.Frame(root, padx=20, pady=10, bg="#f0f0f0")
frame.pack(expand=True)

def criar_label(texto, row):
    label = tk.Label(frame, text=texto, bg="#f0f0f0", font=("Arial", 10, "bold"), anchor="w")
    label.grid(row=row, column=0, sticky="w", pady=(5, 0))
    return label

def criar_entry(var, row):
    entry = tk.Entry(frame, textvariable=var, width=50)
    entry.grid(row=row, column=0, pady=2)
    entry.bind("<Return>", iniciar_divisao)  # Permitir envio com Enter
    return entry

criar_label("Arquivo de entrada:", 0)
entrada_var = tk.StringVar()
criar_entry(entrada_var, 1)
tk.Button(frame, text="Selecionar", command=selecionar_arquivo).grid(row=1, column=1, padx=5)

criar_label("Pasta de saída:", 2)
saida_var = tk.StringVar()
criar_entry(saida_var, 3)
tk.Button(frame, text="Selecionar", command=selecionar_pasta).grid(row=3, column=1, padx=5)

criar_label("Nome base dos arquivos:", 4)
nome_var = tk.StringVar()
criar_entry(nome_var, 5)

criar_label("Quantidade de contatos por arquivo:", 6)
tamanho_lote_var = tk.StringVar()
criar_entry(tamanho_lote_var, 7)

progress_var = tk.IntVar()
progress_bar = ttk.Progressbar(frame, variable=progress_var, maximum=100, length=400, mode="determinate")
progress_bar.grid(row=8, column=0, pady=10, columnspan=2)

tk.Button(frame, text="Iniciar", command=iniciar_divisao, bg="#4CAF50", fg="white", padx=10, pady=5).grid(row=9, column=0, pady=10, columnspan=2)

root.mainloop()
